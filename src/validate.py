#!/usr/bin/env python3
#
# This script reads the NCBI Taxonomy files: names.dmp and nodes.dmp, as well as an Excel (xlsx)
# file, looks for a column named  'Virus Strain' in the excel file, then checks that every value in
# that column is a valid NCBI Taxonomy name. It writes a copy of the Excel file with highlighting
# and comments on match results for each cell.
#
# Download NCBI Taxonomy data from:
# <ftp://ftp.ncbi.nih.gov/pub/taxonomy/taxdmp.zip>
#
# Requirements:
# - Python 3
# - [openpyxl](http://openpyxl.readthedocs.io)

import argparse
import re

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# Configuration
author = 'HIPC Validation Service'
greenFill = PatternFill(start_color='D8FFD8', end_color='D8FFD8', fill_type='solid')
blueFill = PatternFill(start_color='CCD8FF', end_color='CCD8FF', fill_type='solid')
orangeFill = PatternFill(start_color='FFF1D8', end_color='FFF1D8', fill_type='solid')
redFill = PatternFill(start_color='FFD8D8', end_color='FFD8D8', fill_type='solid')
darkRedFill = PatternFill(start_color='FFBBBB', end_color='FFBBBB', fill_type='solid')

# Load NCBI Taxonomy data into various dictionaries
parents = {}
taxid_names = {}
scientific_names = {}
synonyms = {}
lowercase_names = {}


def load_nodes(path):
  """Given a path to the NCBI nodes.dmp file,
  fill the `parents` dictionary."""
  global parents
  with open(path, 'r') as r:
    for line in r:
      (taxid, parent, other) = re.split(r'\s*\|\s*', line.strip('|\n\t '), 2)
      parents[taxid] = parent


def load_names(path):
  """Given a path to the NCBI names.dmp file,
  fill the various names dictionaries."""
  global taxid_names, scientific_names, synonyms, lowercase_names
  with open(path, 'r') as r:
    for line in r:
      (taxid, name, unique, kind) = re.split(r'\s*\|\s*', line.strip('|\n\t '), 3)
      if kind == 'scientific name':
        taxid_names[taxid] = name
        scientific_names[name] = taxid
      else:
        synonyms[name] = taxid
      lowercase_names[name.lower()] = taxid


def is_virus(taxid):
  """Given a taxonomy ID, return true if it is a virus, false otherwise."""
  return taxid and taxid != '1' and (taxid == '10239' or is_virus(parents[taxid]))


def match_taxon(name):
  """Given a name, try to match a taxon,
  and return a tuple of: name, taxid, scientific_name, automatic_replacement"""
  taxid = None
  scientific_name = None
  automatic_replacement = False
  iname = name.strip().lower().replace('  ', ' ')

  if name:
    # 1. 'name' matches the scientific name of a virus:
    if name in scientific_names:
      taxid = scientific_names[name]
      scientific_name = name
    # 2. 'name' is a close case-insensitive match for a virus:
    elif iname in lowercase_names:
      taxid = lowercase_names[iname]
      scientific_name = taxid_names[taxid]
      automatic_replacement = True
    # 3. 'name' is the exact synonym of some taxon
    elif name in synonyms:
      taxid = synonyms[name]
      scientific_name = taxid_names[taxid]
    # 4. 'name' is a substring of exactly one scientific name:
    else:
      matches = []
      for scientific_name in scientific_names.keys():
        if name in scientific_name:
          matches.append(scientific_name)
          if len(matches) > 1:
            break
      if len(matches) == 1:
        scientific_name = matches[0]
        taxid = scientific_names[scientific_name]

  return (name, taxid, scientific_name, automatic_replacement)


def validate_taxon(cell):
  """Given a cell with a taxon value,
  check that it is valid, make a suggestion, or mark it as an error."""
  (name, taxid, scientific_name, automatic_replacement) = match_taxon(cell.value or '')

  # Update the cell
  if is_virus(taxid):
    if name == scientific_name:
      cell.fill = greenFill
    elif automatic_replacement:
      cell.comment = Comment('Automatically replaced "%s" with "%s".'
                             % (name, scientific_name), author)
      cell.value = scientific_name
      cell.fill = blueFill
    else:
      cell.comment = Comment('Suggestion: ' + scientific_name, author)
      cell.fill = orangeFill
    return scientific_name
  elif taxid:
    cell.comment = Comment('Not the name of a virus', author)
    cell.fill = redFill
  else:
    cell.comment = Comment('Not found in NCBI Taxonomy', author)
    cell.fill = darkRedFill


def process_workbook(in_path, out_path):
  """Load an Excel file, search for the 'Taxon Virus Strain' column,
  then validate each virus name."""
  wb = load_workbook(in_path)
  ws = wb.active
  column = None
  results = set()
  for row in ws:
    if column:
      result = validate_taxon(row[column])
      if result:
        results.add(result)
    else:
      for cell in row:
        if cell.value == 'Virus Strain':
          column = cell.col_idx - 1

  # Add suggested values to a lookup column, and replace 'lookupvirus_strain' named range
  ws = wb['lookup']
  values = ['Foo', 'BAR', 'Baz']
  r = 2
  for value in sorted(results):
    ws.cell(row=r, column=2).value = value
    r += 1
  rng = wb.get_named_range('lookupvirus_strain')
  if rng:
    wb.remove_named_range(rng)
    wb.create_named_range('lookupvirus_strain', wb['lookup'], '$B$1:$B$' + str(r))

  wb.save(out_path)


if __name__ == "__main__":
  parser = argparse.ArgumentParser(
    description='Validate taxon names in a spreadsheet. Download NCBI Taxonomy from '
    'ftp://ftp.ncbi.nih.gov/pub/taxonomy/taxdmp.zip')
  parser.add_argument('nodes', type=str, help='The NCBI nodes.dmp file')
  parser.add_argument('names', type=str, help='The NCBI names.dmp file')
  parser.add_argument('input', type=str, help='The XLSX file to read')
  parser.add_argument('output', type=str, help='The XLSX file to write')
  args = parser.parse_args()

  load_nodes(args.nodes)
  load_names(args.names)
  process_workbook(args.input, args.output)


# Unit tests:

def test_match_taxon():
  scientific_names['FOO'] = '1234'
  taxid_names['1234'] = 'FOO'
  synonyms['bAR'] = '1234'
  lowercase_names['foo'] = '1234'
  lowercase_names['bar'] = '1234'

  (name, taxid, scientific_name, automatic_replacement) = match_taxon('FOO')
  assert name == 'FOO'
  assert taxid == '1234'
  assert scientific_name == 'FOO'
  assert automatic_replacement is False

  (name, taxid, scientific_name, automatic_replacement) = match_taxon('  FOO  ')
  assert name == '  FOO  '
  assert taxid == '1234'
  assert scientific_name == 'FOO'
  assert automatic_replacement is True

  (name, taxid, scientific_name, automatic_replacement) = match_taxon('FO')
  assert name == 'FO'
  assert taxid == '1234'
  assert scientific_name == 'FOO'
  assert automatic_replacement is False
